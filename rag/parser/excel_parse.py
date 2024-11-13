from openpyxl import load_workbook
import xlrd
class ExcelParser:
    def xlsx(self, file_path) -> str:
        wb = load_workbook(file_path)
        tb_chunks = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            tb = f"<table><caption>{sheet_name}</caption>"

            # 获取标题行
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
            tb += "<tr>"
            for header_cell in header_row:
                tb += f"<th>{header_cell.value}</th>"
            tb += "</tr>"

            rows = list(sheet.iter_rows(min_row=2, values_only=False))
            for r in rows:
                tb += "<tr>"
                for c in r:
                    cell_value = c.value
                    tb += f"<td>{cell_value if cell_value is not None else ''}</td>"
                tb += "</tr>"

            tb += "</table>"
            tb_chunks.append(tb)
        html_content = '<html><body>' + ''.join(tb_chunks) + '</body></html>'
        return  html_content
    
    def xls(self, file_path) -> str:

        wb = xlrd.open_workbook(file_path)

        html_chunks = []

        # 遍历工作簿中的每一个工作表
        for sheet_index in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_index)
            merged_cells = sheet.merged_cells

            html_table = "<table>"
            html_table += f"<caption>{sheet.name}</caption>"

            # 处理表头
            header_row = sheet.row(0)
            html_table += "<tr>"
            for cell in header_row:
                html_table += f"<th>{cell.value}</th>"
            html_table += "</tr>"

            # 处理数据行
            for row in range(1, sheet.nrows):
                html_table += "<tr>"
                for col in range(sheet.ncols):
                    is_merged = False
                    for merged_cell in merged_cells:
                        if merged_cell.min_row <= row <= merged_cell.max_row and merged_cell.min_col <= col <= merged_cell.max_col:
                            is_merged = True
                            break

                    if is_merged:
                        cell_value = sheet.cell(merged_cell.min_row, merged_cell.min_col).value
                    else:
                        cell_value = sheet.cell(row, col).value

                    html_table += f"<td>{cell_value if cell_value is not None else ''}</td>"
                html_table += "</tr>"

            html_table += "</table>"
            html_chunks.append(html_table)

        html_content = '<html><body>' + ''.join(html_chunks) + '</body></html>'

        # 返回HTML内容而不是文件路径
        return html_content
if __name__ == "__main__":
    
    psr = ExcelParser()
    text = psr.xlsx("test.xlsx")
    print(text)